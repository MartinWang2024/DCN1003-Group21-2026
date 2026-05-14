#!/usr/bin/env python3
"""Import test_data xlsx/xls files into DCN.db (multi-semester, idempotent).

Usage:
    python scripts/import_xlsx.py --db cmake-build-debug/programs/Server/data/DCN.db \
                                   --dir D:\\Handle\\dev\\dcn\\test_data
"""
import argparse
import os
import re
import sqlite3
import sys
from pathlib import Path

# Column layout (row3+ is data; row1 title, row2 header)
COL_CODE = 0
COL_TITLE_SECTION = 1
COL_TEACHERS = 7
COL_SCHEDULE = 8
COL_HOURS = 9
COL_CLASSROOM = 10

# Filename: "Semester N of AYyyyy-yy"
RE_FNAME = re.compile(r"Semester\s+(\d)\s+of\s+AY(\d{4})-(\d{2})", re.IGNORECASE)
# Title cell: "Some Course Title (1001)"
RE_TITLE = re.compile(r"^(.*?)\s*\((\d+)\)\s*$")
# Schedule cell: "Mon 16:00-17:50" possibly with multiple entries separated by ;,/ newline
RE_SCHED = re.compile(r"([A-Za-z]{2,4})\s+(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})")


def parse_semester(filename: str) -> str | None:
    m = RE_FNAME.search(filename)
    if not m:
        return None
    n, yy1, yy2 = m.group(1), m.group(2), m.group(3)
    return f"AY{yy1}-{yy2}_S{n}"


def parse_title_section(cell) -> tuple[str, str]:
    if cell is None:
        return "", ""
    s = str(cell).strip()
    m = RE_TITLE.match(s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return s, ""


def to_int(v, default=0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def to_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def parse_schedules(cell, hours_int: int) -> list[tuple[str, str]]:
    """Return list of (day, duration) tuples. duration encodes 'HH:MM-HH:MM'."""
    if cell is None:
        return []
    text = str(cell)
    results = []
    for m in RE_SCHED.finditer(text):
        day = m.group(1).strip()
        dur = f"{m.group(2)}-{m.group(3)}"
        results.append((day, dur))
    if not results:
        # fallback: store raw with hours
        raw = text.strip()
        if raw:
            results.append((raw[:8], str(hours_int)))
    return results


def read_xlsx_rows(path: Path):
    from openpyxl import load_workbook
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows[2:]  # skip title row + header row


def read_xls_rows(path: Path):
    import xlrd
    book = xlrd.open_workbook(str(path))
    sh = book.sheet_by_index(0)
    return [sh.row_values(r) for r in range(2, sh.nrows)]


def read_rows(path: Path):
    suf = path.suffix.lower()
    if suf == ".xlsx":
        return read_xlsx_rows(path)
    if suf == ".xls":
        return read_xls_rows(path)
    return []


def ensure_schema(conn: sqlite3.Connection):
    cur = conn.cursor()
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS courses (
            code VARCHAR(32) NOT NULL,
            title VARCHAR(128) NOT NULL,
            section VARCHAR(32) NOT NULL,
            instructor VARCHAR(64) NOT NULL,
            PRIMARY KEY(code, section)
        );
        CREATE TABLE IF NOT EXISTS schedules (
            course_code VARCHAR(32) NOT NULL,
            section VARCHAR(32) NOT NULL,
            day VARCHAR(32) NOT NULL,
            duration VARCHAR(32) NOT NULL,
            semester VARCHAR(16) NOT NULL DEFAULT '',
            classroom VARCHAR(64) NOT NULL DEFAULT '',
            PRIMARY KEY(course_code, section, day, duration, semester),
            FOREIGN KEY(course_code, section) REFERENCES courses(code, section)
                ON DELETE CASCADE ON UPDATE CASCADE
        );
    """)
    conn.commit()


def import_file(conn: sqlite3.Connection, path: Path) -> tuple[int, int, int]:
    semester = parse_semester(path.name)
    if not semester:
        print(f"[SKIP] cannot parse semester: {path.name}", file=sys.stderr)
        return (0, 0, 0)

    rows = read_rows(path)
    cur = conn.cursor()
    n_courses = n_sched = n_skipped = 0

    for row in rows:
        if not row:
            n_skipped += 1
            continue
        if len(row) <= COL_CLASSROOM:
            n_skipped += 1
            continue
        code = to_str(row[COL_CODE])
        if not code or code.lower() in ("course code", "none"):
            n_skipped += 1
            continue
        title, section = parse_title_section(row[COL_TITLE_SECTION])
        if not section:
            n_skipped += 1
            continue
        instructor = to_str(row[COL_TEACHERS])
        classroom = to_str(row[COL_CLASSROOM])
        hours = to_int(row[COL_HOURS], 0)
        scheds = parse_schedules(row[COL_SCHEDULE], hours)
        if not scheds:
            n_skipped += 1
            continue

        cur.execute(
            "INSERT OR IGNORE INTO courses(code,title,section,instructor) VALUES(?,?,?,?)",
            (code, title, section, instructor),
        )
        n_courses += cur.rowcount if cur.rowcount > 0 else 0

        for day, dur in scheds:
            cur.execute(
                "INSERT OR REPLACE INTO schedules"
                "(course_code,section,day,duration,semester,classroom)"
                " VALUES(?,?,?,?,?,?)",
                (code, section, day, dur, semester, classroom),
            )
            n_sched += 1

    conn.commit()
    print(f"[OK] {path.name}: semester={semester} courses+={n_courses} schedules={n_sched} skipped={n_skipped}")
    return (n_courses, n_sched, n_skipped)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--dir", required=True)
    ap.add_argument("--pattern", default="Course List and Timetable_*.xls*")
    args = ap.parse_args()

    db_path = Path(args.db)
    src_dir = Path(args.dir)
    db_path.parent.mkdir(parents=True, exist_ok=True)

    files = sorted(src_dir.glob(args.pattern))
    if not files:
        print(f"[ERR] no files match in {src_dir}", file=sys.stderr)
        sys.exit(2)

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys=ON;")
    ensure_schema(conn)

    totals = [0, 0, 0]
    for f in files:
        c, s, k = import_file(conn, f)
        totals[0] += c
        totals[1] += s
        totals[2] += k

    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM courses;")
    course_total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM schedules;")
    sched_total = cur.fetchone()[0]
    cur.execute("SELECT semester, COUNT(*) FROM schedules GROUP BY semester ORDER BY semester;")
    by_sem = cur.fetchall()
    conn.close()

    print("---")
    print(f"Inserted courses(new): {totals[0]}, schedules upserts: {totals[1]}, skipped rows: {totals[2]}")
    print(f"DB totals: courses={course_total}, schedules={sched_total}")
    for sem, cnt in by_sem:
        print(f"  {sem}: {cnt}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Import course Excel files into the DCN course SQLite DB.

Aggregates multi-slot rows per (code, section) into one row:
  day       = "Mon 15:00-16:50; Thu 11:00-11:50"
  duration  = total minutes (sum of "Hours" col * 60)
  classroom = "T6-705; T4-604"

When importing multiple semesters, pass --prefix S2526S1- to disambiguate
sections (PK = code+section) across semesters.

Schema matches Driver/include/database.h CourseRepository.initialize_schema().
"""
from __future__ import annotations

import argparse
import os
import re
import sqlite3
import sys
from collections import OrderedDict
from typing import Iterable, List, Optional, Tuple

HEADER_KEYS = {
    "code": "Course Code",
    "title": "Course Title & Session",
    "teachers": "Teachers",
    "schedule": "Class Schedule",
    "hours": "Hours",
    "classroom": "Classroom",
}

SCHEMA_SQL = (
    "CREATE TABLE IF NOT EXISTS courses ("
    "code TEXT NOT NULL,"
    "title TEXT NOT NULL,"
    "section TEXT NOT NULL,"
    "instructor TEXT NOT NULL,"
    "day TEXT NOT NULL,"
    "duration TEXT NOT NULL,"
    "classroom TEXT NOT NULL,"
    "PRIMARY KEY (code, section)"
    ");"
    "CREATE INDEX IF NOT EXISTS idx_courses_instructor ON courses(instructor);"
)


def iter_excel_rows(path: str) -> Tuple[List[str], Iterable[List]]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header_row = None
        for row in rows:
            cells = [("" if v is None else str(v)).strip() for v in row]
            if HEADER_KEYS["code"] in cells:
                header_row = cells
                break
        if header_row is None:
            raise RuntimeError(f"header row not found in {path}")
        body = ([("" if v is None else v) for v in row] for row in rows)
        return header_row, body
    elif ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        header_idx = None
        header_row: List[str] = []
        for r in range(min(ws.nrows, 5)):
            row = [str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)]
            if HEADER_KEYS["code"] in row:
                header_idx = r
                header_row = row
                break
        if header_idx is None:
            raise RuntimeError(f"header row not found in {path}")
        def gen():
            for r in range(header_idx + 1, ws.nrows):
                yield [ws.cell_value(r, c) for c in range(ws.ncols)]
        return header_row, gen()
    else:
        raise RuntimeError(f"unsupported extension: {ext}")


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


SECTION_RE = re.compile(r"\(([^)]+)\)\s*$")


def split_title_section(title_cell: str) -> Tuple[str, str]:
    s = title_cell.strip()
    m = SECTION_RE.search(s)
    if not m:
        return s, ""
    section = m.group(1).strip()
    title = s[: m.start()].strip()
    return title, section


def hours_to_minutes(h: str) -> int:
    h = h.strip()
    if not h:
        return 0
    try:
        return int(round(float(h) * 60))
    except ValueError:
        return 0


def aggregate_rows(header: List[str], body: Iterable[List]) -> List[dict]:
    col = {k: header.index(v) for k, v in HEADER_KEYS.items() if v in header}
    if "code" not in col or "title" not in col or "schedule" not in col:
        raise RuntimeError(f"required columns missing from header: {header}")

    agg: "OrderedDict[Tuple[str,str], dict]" = OrderedDict()
    for raw in body:
        code = cell_str(raw[col["code"]])
        if not code:
            continue
        title_cell = cell_str(raw[col["title"]])
        title, section = split_title_section(title_cell)
        if not section:
            section = "0000"
        teachers = cell_str(raw[col["teachers"]]) if "teachers" in col else ""
        schedule = cell_str(raw[col["schedule"]]) if "schedule" in col else ""
        hours = cell_str(raw[col["hours"]]) if "hours" in col else ""
        classroom = cell_str(raw[col["classroom"]]) if "classroom" in col else ""

        key = (code, section)
        cur = agg.get(key)
        if cur is None:
            cur = {
                "code": code,
                "title": title,
                "section": section,
                "instructor": teachers,
                "day_parts": [],
                "duration_min": 0,
                "classroom_parts": [],
            }
            agg[key] = cur
        if schedule and schedule not in cur["day_parts"]:
            cur["day_parts"].append(schedule)
        cur["duration_min"] += hours_to_minutes(hours)
        if classroom and classroom not in cur["classroom_parts"]:
            cur["classroom_parts"].append(classroom)
        if teachers and not cur["instructor"]:
            cur["instructor"] = teachers

    out: List[dict] = []
    for v in agg.values():
        out.append({
            "code": v["code"],
            "title": v["title"],
            "section": v["section"],
            "instructor": v["instructor"],
            "day": "; ".join(v["day_parts"]),
            "duration": str(v["duration_min"]),
            "classroom": "; ".join(v["classroom_parts"]),
        })
    return out


def import_file(conn: sqlite3.Connection, path: str, prefix: str,
                clear_first: bool) -> int:
    header, body = iter_excel_rows(path)
    courses = aggregate_rows(header, body)
    if prefix:
        for c in courses:
            c["section"] = f"{prefix}{c['section']}"
    cur = conn.cursor()
    cur.executescript(SCHEMA_SQL)
    if clear_first:
        cur.execute("DELETE FROM courses;")
    cur.executemany(
        "INSERT OR REPLACE INTO courses "
        "(code, title, section, instructor, day, duration, classroom) "
        "VALUES (?, ?, ?, ?, ?, ?, ?);",
        [
            (c["code"], c["title"], c["section"], c["instructor"],
             c["day"], c["duration"], c["classroom"])
            for c in courses
        ],
    )
    conn.commit()
    return len(courses)


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("inputs", nargs="+", help="Excel files (.xlsx or .xls)")
    ap.add_argument("--db", default="data/courses.db",
                    help="SQLite DB path (default: data/courses.db)")
    ap.add_argument("--prefix", default="",
                    help="Section prefix to disambiguate semesters, "
                         "e.g. S2526S1-")
    ap.add_argument("--clear", action="store_true",
                    help="DELETE FROM courses before first import")
    args = ap.parse_args(argv)

    os.makedirs(os.path.dirname(os.path.abspath(args.db)) or ".",
                exist_ok=True)
    conn = sqlite3.connect(args.db)
    conn.execute("PRAGMA journal_mode=WAL;")
    total = 0
    for i, p in enumerate(args.inputs):
        n = import_file(conn, p, args.prefix,
                        clear_first=(args.clear and i == 0))
        print(f"  + {n:5d} rows from {os.path.basename(p)}")
        total += n
    cur = conn.execute("SELECT COUNT(*) FROM courses;")
    db_count = cur.fetchone()[0]
    print(f"imported {total} aggregated row(s); DB now holds {db_count}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
